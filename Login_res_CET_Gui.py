import os
import time
import socket
import openpyxl
#import random
import PySimpleGUI as sg
sg.theme('SandyBeach')
import requests
import re

os.chdir(os.path.dirname(__file__))
file_path = os.getcwd()

print(file_path)

# imph = //*[@id="ctl00_MainContent_txt_Captcha"]

CURRENT_YEAR = time.strftime("%Y")
print(CURRENT_YEAR)


from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager




VIRTUAL_DISP = True
VIRTUAL_DISP = False

BROWSER_NAME = 'FIREFOX'
BROWSER_NAME = 'CHROME'
#BROWSER_NAME = 'INTERNET EXPLORER'

QUIT_BROWSER = True
QUIT_BROWSER = False
 
EXP_TIME = 1800



WAIT_TIME = 45


RANK = ''
ROLL = ''

CET_YEAR = {"EAMCETB":f"{int(CURRENT_YEAR)-2}",
            "EAMCET":f"{int(CURRENT_YEAR)-2}",
            "ICET": f"{int(CURRENT_YEAR)-5}",
            "ECET" : f"{int(CURRENT_YEAR)-3}",
            "POLYCET" : f"{int(CURRENT_YEAR)}" ,
            "DPHARM": f"{int(CURRENT_YEAR)-2}"
            }


CET_MAP = {"EAMCETB":"155",
        "EAMCET":"199",
            "ICET": "166",
            "ECET" : "188",
            "POLYCET" : "177",
           "DPHARM":"122"
            }

CET_NAME = ["EAMCETB","EAMCET","ICET","ECET","POLYCET","DPHARM"]
while(True):
        try:
                i=0
                for key in CET_NAME:
                        i+=1
                        print(f'{i}.{key}')
                key = int(input('Enter set number : '))
                print(CET_NAME[key-1]+''+CET_MAP[CET_NAME[key-1]])
                break
        except:
                print("wrong Number entered try again ")

drives = [ chr(x) + ":" for x in range(65,91) if os.path.exists(chr(x) + ":") ]

path = f'{drives[-1]}/COUNSELING_DATA/{CET_NAME[key-1]}_{CURRENT_YEAR}/{time.strftime("%d-%B-%Y")}/'
try:
        os.makedirs(path)
except OSError:
        if not os.path.isdir(path):
                raise


        

LOGIN_TYPE = {
                "V":"V",
                "R":"R",
                "H":"H"        }

LOGIN_NAME =  ['H',"R","V"]
while(True):
        try:
                i=0
                os.system('cls')
                for l in LOGIN_NAME:
                        if l != 'H' :
                                i=i+1
                                print(f'{i}.{l}')
                l = int(input('Enter login type number : '))
                print(LOGIN_NAME[l]+' '+LOGIN_TYPE[LOGIN_NAME[l]])
                break
        except:
                print("wrong Number entered try again ")


if LOGIN_NAME[l] == 'R':
        while(True):
                try:
                        os.system('cls')
                        print("\n\n")
                        vo = int(input('Enter Number of Verification Officers(min 1 and Max 6) : '))
                        if vo>0 and vo<7:
                                
                                print(f"Number of Verification Officers: {vo}")
                                VerOff_num = vo
                                break
                except:
                        print("wrong Number entered try again")
SSC_LEN = len("2118117165")
CBSE_LEN = len("20158163")
MAHA_LEN = len("J151930")
MOBILE_LEN = len("9963459392")

if CET_NAME[key-1] ==   "POLYCET":
        MAX_RANK = 100000
        HALL_lEN = len("1234567")
elif CET_NAME[key-1] ==   "ECET":
        MAX_RANK = 28000
        HALL_lEN = len("12345678901")
elif CET_NAME[key-1] ==   "EAMCET":
        MAX_RANK = 250000
        HALL_lEN = len("2025C07164")
elif CET_NAME[key-1] ==   "EAMCETB":
        MAX_RANK = 100000
        HALL_lEN = len("2025C07164")
elif CET_NAME[key-1] ==   "ICET":
        MAX_RANK = 100000
        HALL_lEN = len("2026302193")
elif CET_NAME[key-1] ==   "DPHARM":
        MAX_RANK = 100000
        HALL_lEN = 11


else:
        MAX_RANK ="NIL"
print(f"MAX RANK for {CET_NAME[key-1]}: {MAX_RANK}")
                


CET_PORTAL_URL = f'https://ts{CET_NAME[key-1]}d.nic.in/'

TOTAL = 0
work =0
path_Attempts = f'{path}{LOGIN_NAME[l]}.xlsx'




if VIRTUAL_DISP:
        print("starting virtual diplay")
        display = Display(visible=0, size=(800, 600))
        display.start()
        print('virtual dispaly started')
else:
        #print('No virtual Display, diplaying on monitor')
        print("starting........")
def dip_res(PIN):
        r = requests.get(f'https://www.sbtet.telangana.gov.in/API/api/Results/GetConsolidatedResults?Pin={PIN}')
        r = r.text
        #print(r)
        if ('CGPA' in r):
                y = r.split('CGPA')
                #print(y)
                y = y[1].split(':')
                #print(y)
                y = y[1].split(',')
                y = y[0]
                #print(y)
                y = [float(s) for s in re.findall(r'-?\d+\.?\d*', y)]
                if len(y)==0:
                        r = requests.get(f'https://www.sbtet.telangana.gov.in/API/api/PreExamination/getDetailsByPins?pin={PIN}')
                        r = r.text
                        #print(r)
                        if ('\\"CGPA\\":\\"' in r):
                                y = r.split('\\"CGPA\\":\\"')
                                y = y[1].split('\\",')
                                y = y[0]
                                return f',{PIN}, CGPA:{y}, %: {(float(y)-0.5)*10:.2f}'
                        if ('PercentageOfMarks' in r):
                                y = r.split('PercentageOfMarks')
                                y = y[1].split(':')
                                y = y[1].split(',')
                                y = y[0]
                                y = [float(s) for s in re.findall(r'-?\d+\.?\d*', y)]
                                
                                return f', {PIN}, PER :{y[0]}'
                        else:
                            return ''
                        
                else:                    
                    return f', {PIN}, CGPA:{y[0]}, %: {(float(y[0])-0.5)*10:.2f}'
        
        else:
                return ''
                
        

def get_captcha(driver):

    # now that we have the preliminary stuff out of the way time to get that image :D
    element = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_MainContent_Panel1"]/table/tbody/tr[7]/td[1]/div/img')))
    
    location = element.location
    size = element.size
    
    # saves screenshot of entire page
    driver.save_screenshot('screenshot.png')

    # uses PIL library to open image in memory
    image = Image.open('screenshot.png')
    
    left = location['x']
    top = location['y']
    right = location['x'] + size['width']
    bottom = location['y'] + size['height']
    
    image = image.crop((left, top, right, bottom))  # defines crop points
    print("yes-p")
    try:
            image.save('screenshot.png')  # saves new cropped image
    except Exception as e:
            print(e.message)
    #ptint("yes")
    image = cv2.imread('screenshot.png')
    image[:,:,2] = np.zeros([image.shape[0], image.shape[1]])
    image = cv2.resize(image, None, fx=2, fy=4, interpolation=cv2.INTER_CUBIC)
    kernel = np.ones((1, 1), np.uint8)
    image = cv2.dilate(image, kernel, iterations=1)
    image = cv2.erode(image, kernel, iterations=1)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (3,3), 0)
    thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]

    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe'
    data = pytesseract.image_to_string(thresh, lang='eng', config='--psm 6')
    data = data.split('\n')[0]
    data = data.replace(' ','').upper()
    print(data)
    return data
        
        
def init_driver():
        print('Opening Browser........')
        
        if BROWSER_NAME == 'CHROME':
                from selenium.webdriver.chrome.service import Service
                chrome_profile = webdriver.ChromeOptions()
                chrome_profile.add_argument('--kiosk-printing')
                chrome_profile.add_argument('--start-maximized')
                chrome_profile.add_argument('--disable-infobars')
                print('Opening {0} Browser........'.format(BROWSER_NAME))
                driver = webdriver.Chrome(service =Service(ChromeDriverManager().install()),options=chrome_profile)
                #driver = webdriver.Firefox()
        if BROWSER_NAME == 'FIREFOX':
                print('Opening {0} Browser........'.format(BROWSER_NAME))
                driver = webdriver.Firefox()
        if BROWSER_NAME == 'INTERNET EXPLORER':
                print('Opening {0} Browser........'.format(BROWSER_NAME))
                driver = webdriver.Ie()
                #driver = webdriver.Firefox()
                
        driver.wait = WebDriverWait(driver, EXP_TIME)
        print('{0} Browser opened'.format(BROWSER_NAME))
        return driver

def init_Portal(driver):
        print("Opening {0}".format(CET_PORTAL_URL))
        driver.get(CET_PORTAL_URL)
        print('{0} opened'.format(CET_PORTAL_URL))
        

        
def quit_driver(driver):
        
        if QUIT_BROWSER:
                print("Quiting Browser......")
                driver.quit()
                print(" Browser Closed . Bye!Bye!")
        else:
                print("NOT Quiting Browser......")
                print("Please close Browser Manually......")
                
def login(driver,lid,pwd,COLL_CODE,login_Type):
        
        print('loging in........')
        #box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="popup"]/img')))
        #box.click()
        box = driver.wait.until(EC.presence_of_element_located((By.ID, 'ctl00_MainContent_user_type')))
        #box.clear()
        mySelect = Select(box)
        mySelect.select_by_value(login_Type)
        time.sleep(1)
        
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_MainContent_h123"]')))
        box.clear()
        box.send_keys(lid)
        #button = driver.wait.until(EC.element_to_be_clickable((By.NAME, "btnsubmit")))
        #time.sleep(30)
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_MainContent_pqrabc"]')))
        box.clear()
        box.send_keys(pwd)
        #button = driver.wait.until(EC.element_to_be_clickable((By.NAME, "btnsubmit")))

        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_MainContent_ccode"]')))
        box.clear()
        box.send_keys(COLL_CODE)

        # image = //*[@id="ctl00_MainContent_Panel1"]/table/tbody/tr[7]/td[1]/div/img
        #button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_MainContent_Panel1"]/table/tbody/tr[7]/td[1]/div/img')))
        #captcha_text = get_captcha()
        box = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_MainContent_txt_Captcha"]')))
        box.clear()
        box.click()
        #box.send_keys(get_captcha(driver))
        
        time.sleep(5)
        print('Succesfully logged in with USERNAME : {0}........'.format(lid))
        
def get_Roll_Rank(driver):
        ROLL = "ACIDONE"
        RANK= "ACIDONE"
        name_dsip= "ACIDONE"
        fath_dsip="ACIDONE"
        aci_done = False
        layout = [   
                                                                #[sg.Text(f' Welcome to  Registration ',auto_size_text= True, font=('Helvetica', 50))],
                                                                [sg.Text(f'Press ENTER Button to Start',auto_size_text= True, font=('Helvetica', 75))],
                                                                [sg.Submit(font=('Helvetica', 42))]] 
        window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
        window.Maximize()
        event, values = window.read()
        warn = ""
        dip_result= ''
        reenter = False
        while True:

                if CET_NAME[key-1] ==   "ECET":
                        ecount = 0
                        while True:
                    
                            layout = [ [sg.Text(f'{warn}',auto_size_text= True, font=('Helvetica', 42))], 
                            [sg.Text(f'Please enter your DIPLOMA PIN Number:',auto_size_text= True, font=('Helvetica', 35))], 
                            [sg.Text(f'DIPLOMA PIN Number:',auto_size_text= True, font=('Helvetica', 42)), sg.InputText(font=('Helvetica', 42))],
                            [sg.Text(f'Press ENTER Button to Submit',auto_size_text= True, font=('Helvetica', 42))], 
                            [sg.Submit(font=('Helvetica', 42))] 
                            ]
                            window.close()
                            window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                            window.Maximize()
                            event, values = window.read()
                    #print(event, values[0])
                            values[0] = values[0].upper()
                            values[0] = values[0].replace(" ",'')
                            pin_vals = values[0].split("-")
                            if (len(pin_vals) == 3 and len(pin_vals[0]) == 5 and pin_vals[2].isnumeric())   :
                                dip_result = dip_res(values[0])
                                print(dip_result)                       
                                break
                            else:
                                if(ecount<5):
                                        warn = f"ENTERED PIN {values[0]} WRONG! Enter again "
                                else:
                                        warn = f"ENTERED PIN {values[0]} WRONG! Contact Admin"
                                        
                                ecount +=1
                                window.close()
                                

                warn = ""
                        
                
                while True:
                    
                    layout = [ [sg.Text(f'{warn}',auto_size_text= True, font=('Helvetica', 42))], 
                            [sg.Text(f'Please enter your',auto_size_text= True, font=('Helvetica', 35))], 
                            [sg.Text(f'{CET_NAME[key-1]} Hall Ticket Number: ',auto_size_text= True, font=('Helvetica', 42)), sg.InputText(font=('Helvetica', 42))],
                            [sg.Text(f'Press ENTER Button to Submit',auto_size_text= True, font=('Helvetica', 42))], 
                            [sg.Submit(font=('Helvetica', 42))] 
                    ]
                    window.close()
                    window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                    window.Maximize()
                    event, values = window.read()
                    #print(event, values[0])
                    values[0] = values[0].replace(" ",'')
                    if (True if "EAMCET" in CET_NAME[key-1] else values[0].isnumeric())and len(values[0]) ==HALL_lEN :
                        break
                    else:
                        warn = "ENTERED DETAILS WRONG! Enter again"
                        window.close()
                
                
                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_roll_no"]')))
                box.clear()
                box.send_keys(values[0].upper())

                warn = ""
                window.close()
                while True:
                    
                    if CET_NAME[key-1] ==   "POLYCET":
                            ty = " ENGINEERING(M.P.C)"
                    elif CET_NAME[key-1] ==   "ECET":
                            ty = " BRANCH"
                  
                    else:
                            ty =""
                    layout = [ [sg.Text(f'{warn}',auto_size_text= True, font=('Helvetica', 42))], 
                                [sg.Text(f'Please enter your',auto_size_text= True, font=('Helvetica', 35))], 
                                [sg.Text(f'{CET_NAME[key-1]}{ty} RANK: ',auto_size_text= True, font=('Helvetica', 42)), sg.InputText(font=('Helvetica', 42))],
                                [sg.Text(f'Press ENTER Button to Submit',auto_size_text= True, font=('Helvetica', 42))], 
                                [sg.Submit(font=('Helvetica', 42))]] 
                    window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                    window.Maximize()
                    event, values = window.read()
                    #print(event, values[0])
                    values[0] = values[0].replace(" ",'')
                    if  (values[0].split('.')[0].isnumeric())and len(values[0].split('.')[0]) <= len(str(MAX_RANK)):
                        if int(values[0].split('.')[0]) <MAX_RANK:
                                break
                        else:
                                warn = "ENTERED DETAILS WRONG! Enter again"
                                window.close()
                    else:
                        warn = "ENTERED DETAILS WRONG! Enter again"
                        window.close()
                
                layout = [      
                                [sg.Text('Please Wait.........:',auto_size_text= True, font=('Helvetica', 42)) ] ]
                window.close()
                window1 = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                window1.Maximize()
                dum_event, dum_values = window1.Read(timeout=0)
                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_rank"]')))
                box.clear()
                box.send_keys(values[0])
                time.sleep(2)
                
                #button = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Panel1"]/table/tbody/tr[1]/td')))
                #button.click()
                #print("ENTERED")
                button = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="aspnetForm"]')))
                button.submit()
                #print("SUBMITTED")
                time.sleep(1)
                button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Button1"]')))
                button.click()
                if reenter:
                        print("REENT")
                        button = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="aspnetForm"]')))
                        button.submit()
                #print("SUBMITTED")
                        time.sleep(1)
                        #button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Button1"]')))
                        #button.click()
                        
                
                try:
                        element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_lbl_rank"]/b')))
                        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_lbl_rank"]/b')))
                        RANK = box.text
                        if RANK == values[0]:
                                break
                        else:
                                warn = "ENTERED DETAILS WRONG! Enter again"
                                window1.close()
                except:
                        try:
                                element = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_errdis"]/font')))
                                element = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_rank"]')))
                                warn = "ENTERED DETAILS WRONG! Enter again"
                                window1.close()
                                reenter = True
                                
                        except:
                                aci_done = True
                                element = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_errdis"]/font')))

                                layout = [   
                                                                [sg.Text(f' Already Registered ! ',auto_size_text= True, font=('Helvetica', 42))],
                                                                [sg.Text(f' Go to Your Verification Officer',auto_size_text= True, font=('Helvetica', 42))],
                                                                #[sg.Text('        \u2794',auto_size_text= True, font=('Helvetica', 250))]]
                                                                [sg.Text('        \u2190',auto_size_text= True, font=('Helvetica', 250))]] 
                                window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                                window.Maximize()
                                dum_event, dum_values = window.Read(timeout=0)
                                time.sleep(10)
                                window.close()
                                
                                #print("Hello")
                                break
                                
        if aci_done:
                window1.close()
                #print("Hello1")
                return ROLL,RANK,name_dsip,fath_dsip,aci_done
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_lbl_roll"]/b')))
        ROLL = box.text
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_cand_name"]/b')))
        name_dsip = box.text
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_father_name"]/b')))
        fath_dsip = box.text
        
        #print(RANK)
        
        #button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_consent"]')))
        #button.click()
        #box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtLAadharNo"]')))
        #ad=box.get_attribute('value')
        
        #print(ad)
        #if(len(ad) != 12):
                #box.clear()
        
        #button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtLAadharNo"]')))
        #button.click()
        #box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Label1"]')))
        window1.close()
        #print("Hello2")
        return ROLL,RANK,name_dsip,fath_dsip,aci_done,dip_result

        

        

        
        
        

       
       

        

        

def print_Form(driver,roll , rank,TOTAL,dip_result):

        
        print("ENTERD")

        button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_HL_Print_Certificate"]')))
        button.click()
        js_scrpit1 =f'''var HTB = document.getElementById("divTag");
                        var HTab = HTB.getElementsByTagName("table")[0].rows[0];
                        var cell = HTab.insertCell(-1);
                        cell.innerText = "{socket.gethostname()}";
                        cell.align ="right";
                        cell.style = "font-size: 10px; font-family: Arial";
                        var cell = HTab.insertCell(-1);
                        cell.innerText = "COUNTER NO:{(TOTAL%VerOff_num)+1}";
                        cell.align ="right";
                        cell.style = "font-size: 40px; font-family: Arial";'''
        js_scrpit2 =f'''var HTB = document.getElementById("divTag");
                        var cell = HTB.getElementsByTagName("table")[1].rows[7].cells[1];
                        cell.innerText += "    {dip_result}";'''
        
       
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_roll_no"]')))
        box.clear()
        box.send_keys(roll)
       
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_rank"]')))
        box.clear()
        box.send_keys(rank)
        button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Button1"]')))
        
        button.click()
        driver.execute_script(js_scrpit1)
        
        
        #box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_mobileno"]')))
        #mob = box.text
        #box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_ccname"]')))
        #nam = box.text

        



        if CET_NAME[key-1] == 'EAMCET':
                button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Panel2"]/p/input')))
                button.click()
        if CET_NAME[key-1] == 'ECET':
                driver.execute_script(js_scrpit2)
                button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Panel2"]/div[2]/p/input')))
                button.click()
        if CET_NAME[key-1] == 'EAMCETB':
                button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Panel2"]/p/input')))
                button.click()
        if CET_NAME[key-1] == 'POLYCET':
                button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Panel2"]/p/input')))
                button.click()
        if CET_NAME[key-1] == 'ICET':
                #button12 = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Panel2"]/p/input')))
                #button12.click()
                print("Printed")
        print("Form Printed\n")
        #print('')
        #rowNo= rowNo+1
        #sheet.cell(row=rowNo, column=1).value = nam
        #sheet.cell(row=rowNo, column=2).value = rank
        #sheet.cell(row=rowNo, column=3).value = mob

def reg(TOTAL,driver):
                print("Opening ACI Registration.....")
                menu = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_menuBarn1"]/table/tbody/tr/td[1]/a')))
                layout = [ [sg.Text('Please Wait.....',auto_size_text= True, font=('Helvetica', 42))]] 
                window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                window.Maximize()
                dum_event, dum_values = window.Read(timeout=0)
                actions = ActionChains(driver)
                
                actions.move_to_element(menu).perform()

                button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_menuBarn4"]/td/table/tbody/tr/td/a')))
                button.click()
                driver.execute_script("document.body.style.zoom='150%'")
                window.close()
                ROLL,RANK,name_dsip,fath_dsip,aci_done,dip_result = get_Roll_Rank(driver)
                print('H.No.: {0}   Rank : {1}'.format(ROLL,RANK))
                if aci_done:
                        name_dsip = ''
                        return TOTAL,ROLL,RANK,name_dsip
            
                layout = [ [sg.Text('Please Wait.....',auto_size_text= True, font=('Helvetica', 42))]] 
                window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                window.Maximize()
                dum_event, dum_values = window.Read(timeout=0)
                while(True):
                        try:
                                
                                #driver.wait.until(EC.alert_is_present())
                                #alert = driver.switch_to.alert
                                #A = alert.text
                                #print(A)
                                #alert.accept()
                                A = 'Capture done, now please verify.'
                                #print('alert accepted')
                                if (A == 'Capture done, now please verify.' ):
                                        try:
                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_sscrno"]')))
                                                ssc_htno=box.get_attribute('value')
                                                if(len(ssc_htno) != SSC_LEN ):
                                                        #print(f"Len = {ssc_htno}")
                                                        #box.clear()

                                                        warn = ""
                                                        window.close()
                                                        while True:
                                                            
                                                            layout = [ [sg.Text(f'{warn}',auto_size_text= True, font=('Helvetica', 42))], 
                                                                       [sg.Text('Please Enter Your SSC/CBSE HallTicket Number',auto_size_text= True, font=('Helvetica', 42))],
                                                                       [sg.Text(f' NAME:{name_dsip} ',auto_size_text= True, font=('Helvetica', 42))],
                                                                        #[sg.Text(f' Father NAME: {fath_dsip}',auto_size_text= True, font=('Helvetica', 42))],
                                                                       [sg.Text(' SSC/CBSE HallTicket Number: ',auto_size_text= True, font=('Helvetica', 42)), sg.InputText(f'{ssc_htno}',font=('Helvetica', 42))],
                                                                       [sg.Text(f'Press ENTER Button to Submit',auto_size_text= True, font=('Helvetica', 42))], 
                                                                       [sg.Submit(font=('Helvetica', 42))]] 
                                                            window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                                                            window.Maximize()
                                                            event, values = window.read()
                                                            #print(event, values[0])
                                                            #print(values[0])
                                                            values[0] = values[0].replace(" ",'')
                                                            if values[0].isnumeric()and (len(values[0]) == SSC_LEN or len(values[0]) == CBSE_LEN):
                                                                    #print(f"{values[0]} HE")
                                                                    break
                                                            elif len(values[0]) == MAHA_LEN:
                                                                    break
                                                                    
                                                            else:
                                                                warn = "ENTERED DETAILS WRONG! Enter again"
                                                                window.close()
                                                        if ssc_htno != values[0]:
                                                                #print(f"{values[0]} HE")
                                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_sscrno"]')))
                                                                #print("Hello")
                                                                #print(f"Len = {ssc_htno}")
                                                                box.clear()
                                                                box.send_keys(values[0])
                                                                ssc_htno = values[0]
                                                #print(f"Len = {ssc_htno}")
                                                #print("Mob")        
                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_mobile_no"]')))
                                                mo=box.get_attribute('value')
                                                if(len(mo) != 10):
                                                        box.clear()

                                                warn = ""
                                                window.close()
                                                while True:
                                                    
                                                    layout = [ [sg.Text(f'{warn}',auto_size_text= True, font=('Helvetica', 42))], 
                                                               [sg.Text('Please Check Your Mobile Number',auto_size_text= True, font=('Helvetica', 42))],
                                                               [sg.Text(f' NAME:{name_dsip} ',auto_size_text= True, font=('Helvetica', 42))],
                                                                #[sg.Text(f' Father NAME: {fath_dsip}',auto_size_text= True, font=('Helvetica', 42))],
                                                               [sg.Text(' Mobile: ',auto_size_text= True, font=('Helvetica', 42)), sg.InputText(f'{mo}',font=('Helvetica', 42))],
                                                               [sg.Text(f'Press ENTER Button to Submit',auto_size_text= True, font=('Helvetica', 42))], 
	                                                       [sg.Submit(font=('Helvetica', 42))]] 
                                                    window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                                                    window.Maximize()
                                                    event, values = window.read()
                                                    #print(event, values[0])
                                                    values[0] = values[0].replace(" ",'')
                                                    if values[0].isnumeric()and len(values[0]) == MOBILE_LEN:
                                                        break
                                                    else:
                                                        warn = "ENTERED DETAILS WRONG! Enter again"
                                                        window.close()
                                                if mo != values[0]:
                                                        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_mobile_no"]')))
                                                        box.clear()
                                                        box.send_keys(values[0])
                                                        mo = values[0]
                                                        
                                                window.close()
                                                #box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btn_Verify_V2"]')))
                                                #box.click()
                                                #print("Adhaar Verified.")
                                                layout = [ [sg.Text('Please Wait.....',auto_size_text= True, font=('Helvetica', 42))]] 
                                                window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                                                window.Maximize()
                                                dum_event, dum_values = window.Read(timeout=0)
                                                boxC = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtCasteNo"]')))
                                                boxC.clear()
                                                m=boxC.get_attribute('value')
                                                #print(len(m))
                                                #print(m)
                                                boxI = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtIncomeNo"]')))
                                                boxI.clear()
                                                n=boxI.get_attribute('value')

                                                boxE = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtEWSNo"]')))
                                                boxE.clear()
                                                EWS=boxE.get_attribute('value')
                                                print(len(m))
                                                print(m)
                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_sscpassdmth"]')))
                                                box.clear()
                                                box.send_keys("03")
                                                if CET_NAME[key-1] ==   "POLYCET" and len(ssc_htno) == SSC_LEN:
                                                        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_sscpassdyr"]')))
                                                        box.clear()
                                                        box.send_keys(f'20{ssc_htno[0:2]}')
                                                else:
                                                        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_sscpassdyr"]')))
                                                        box.clear()
                                                        box.send_keys(CET_YEAR[CET_NAME[key-1]])
                                                
                                                print(f'Caste Certificate No:{m if(m) else "Details Not Entered by Canditate"}, Income Certificate No:{n if(n) else "Details Not Entered by Canditate"}')
                                                                                                
                                                #box.click()
                                                if(len(m)==0 and len(n)==0):
                                                        print("Caste and Income Details Not Entered by Canditate.")
                                                        print("Skipping Mee Seva Details...")
                                                elif(len(m)!=0 or len(n)!=0):
                                                        print("Getting Details from Mee Seeva....")
                                                        
                                                        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btn_MeeSevaService"]')))
                                                        box.click()
                                                        if(len(m)!=0):
                                                                print("Caste Certificate Details....",end="")
                                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_gv_MeeSeva_Data_Caste"]/tbody/tr[2]/td[1]')))
                                                                print(f'Name:{box.text}',end=', ')
                                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_gv_MeeSeva_Data_Caste"]/tbody/tr[2]/td[4]')))
                                                                print(f'Caste:{box.text}')
                                                                len(box.text)
                                                                if (box.text == "NOT AVAILABLE" or box.text == '' or box.text == "OC" or box.text == "EBC" or box.text == "0BC" or len(box.text) > 3 and len(box.text) < 2):
                                                                        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btn_Caste_Incorrect"]')))
                                                                        box.click()
                                                                        
                                                                else:
                                                                        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btn_Caste_Confirm"]')))
                                                                        box.click()
                                                        if(len(n)!=0):
                                                                print("Income Certificate Details....",end="")
                                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_gv_MeeSeva_Data_Income"]/tbody/tr[2]/td[1]')))
                                                                print(f'Father Name:{box.text}',end=", ")
                                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_gv_MeeSeva_Data_Income"]/tbody/tr[2]/td[2]')))
                                                                print(f'Candidate Name:{box.text}',end=", ")
                                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_gv_MeeSeva_Data_Income"]/tbody/tr[2]/td[3]')))
                                                                print(f'Income:{box.text}')
                                                                if box.text == "NOT AVAILABLE":
                                                                        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btn_Income_Incorrect"]')))
                                                                        box.click()
                                                                else:
                                                                        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btn_Income_Confirm"]')))
                                                                        box.click()


                                                
                                                
                                                
                                                
                                                #box.click()
                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_sscpassdmth"]')))
                                                box.clear()
                                                box.send_keys("04")
                                                
                                                
                                                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                                #driver.execute_script("document.body.style.zoom='150%'")
                                                #box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_sscrno"]')))
                                                #box.click()
                                                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                                
                                                if CET_NAME[key-1] ==   "POLYCET" and len(ssc_htno) == SSC_LEN:
                                                        button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btn_sscservice"]')))
                                                        button.click()
                                                        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_gv_ssc_result"]/tbody/tr[2]/td[1]')))
                                                        if ssc_htno == box.text:
                                                                button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btn_ssc_correct"]')))
                                                                button.click()
                                                        else:
                                                                button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btn_ssc_incorrect"]')))
                                                                button.click()
                                                                
                                                                
                                                                
                                                        
                                                warn = ""
                                                window.close()
                                                while True:
                                                    
                                                    layout = [  [sg.Text('Please Check Your Details',auto_size_text= True, font=('Helvetica', 42))], 
                                                                [sg.Text(f' NAME:{name_dsip} ',auto_size_text= True, font=('Helvetica', 42))],
                                                                [sg.Text(f' Father NAME: {fath_dsip}',auto_size_text= True, font=('Helvetica', 42))],
                                                                [sg.Text(f' Mobile:{mo} ',auto_size_text= True, font=('Helvetica', 42))],
                                                                [sg.Text(f'Press ENTER Button to Submit',auto_size_text= True, font=('Helvetica', 42))], 
                                                                [sg.Submit(font=('Helvetica', 42))]] 
                                                    window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                                                    window.Maximize()
                                                    event, values = window.read()
                                                    break
                                                    
                                                window.close()
                                                

                                                print("Saved Sucessfully")
                                                layout = [   
                                                                #[sg.Text(f' Registered Sucessfully! ',auto_size_text= True, font=('Helvetica', 27))],
                                                                #[sg.Text(f' GO to Verification Block',auto_size_text= True, font=('Helvetica', 45))],
                                                                [sg.Text(f' Go to',auto_size_text= True, font=('Helvetica', 55))],
                                                                #[sg.Text(f' Verification',auto_size_text= True, font=('Helvetica', 60))],
                                                                
                                                                [sg.Text(f' Counter-{((TOTAL+1)%VerOff_num)+1}' if CET_NAME[key-1] !=   "ICET" else '' ,auto_size_text= True, font=('Helvetica', 75))],
                                                                [sg.Text(f' and Sit. We will call again!',auto_size_text= True, font=('Helvetica', 40))],
                                                                [sg.Text(f' Do Not Press Any Button!',auto_size_text= True, font=('Helvetica', 60))],
                                                                #[sg.Text('   \u2794',auto_size_text= True, font=('Helvetica', 200))]]
                                                                [sg.Text('   \u2190',auto_size_text= True, font=('Helvetica', 200))]] 
                                                window = sg.Window('CET REG data entry window', layout,keep_on_top=True).Finalize()
                                                window.Maximize()
                                                dum_event, dum_values = window.Read(timeout=0)
                                               # //*[@id="aspnetForm"]
                                                button = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="aspnetForm"]')))
                                                button.submit()
                                                
                                                #//*[@id="ctl00_ContentPlaceHolder1_btn_Save"]
                                                button = driver.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btn_Save"]')))
                                                button.click()
                                                
                                                TOTAL +=1
                                                
                                                print_Form(driver,ROLL,RANK,TOTAL,dip_result)
                                                time.sleep(1)
                                                window.close()
                                                break
                                        except Exception as e:
                                                driver.wait.until(EC.alert_is_present())
                                                alert = driver.switch_to.alert
                                                A = alert.text
                                                print(A)
                                                alert.accept()
                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtLAadharNo"]')))
                                                box.clear()
                                                box.send_keys(A)
                                                box.click()
                                                box.click()
                                                
                                                
                                

                        except Exception as e:
                                print(e.message)
                                print('H.No.: {0}   Rank : {1}'.format(ROLL,RANK))
                return TOTAL,ROLL,RANK,name_dsip
                        
def ver(TOTAL,driver):
        actions = ActionChains(driver)
        menu = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_menuBarn1"]/table/tbody/tr/td[1]/a')))
        actions.move_to_element(menu).perform()
        if CET_NAME[key-1] != 'EAMCET':
                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_menuBarn4"]/td/table/tbody/tr/td/a')))
                box.click()
        if CET_NAME[key-1] == 'EAMCET':
                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_menuBarn5"]/td/table/tbody/tr/td/a')))
                box.click()
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_roll_no"]')))
        box.click()
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Label6"]')))
        print("getting Details......")
                                                
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_cname"]')))
        print(f"Cadidate Name: {box.text}", end= ", ")
                                                
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_hlt_no"]')))
        print(f"Hall Ticket No: {box.text}", end= ", ")

        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_crank"]')))
        print(f"Rank: {box.text}")
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_eligible"]')))
        #box.clear()
        mySelect = Select(box)
        mySelect.select_by_value('B')
        time.sleep(1)
        box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Label6"]')))
        box.click()
        while(True):
                                try:
                                        
                                        #print('H.No. :{0}   Rank : {1}'.format(ROLL,RANK))
                                        driver.wait.until(EC.alert_is_present())
                                        alert = driver.switch_to.alert
                                        A = alert.text
                                        print(A)
                                        alert.accept()
                                        #print('alert accepted')
                                        if ('Candidate details Successfully updated' in A ):
                                                TOTAL +=1
                                                print("Printing Form......")
                                                
                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_ccname"]')))
                                                print(f"Cadidate Name: {box.text}", end= ", ")
                                                
                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_rollnum"]')))
                                                print(f"Hall Ticket No: {box.text}", end= ", ")

                                                box = driver.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_candrank"]')))
                                                print(f"Hall Ticket No: {box.text}")
                                                print("Form Printed...")
                                                print("")

                                                break
                                except Exception as e:
                                        print(e.message)
        return TOTAL
                                        
                                         
                                                
        

        
try:        
        if __name__ == "__main__":
                #wb = openpyxl.Workbook()
                #wb.create_sheet(index=0, title=CET_NAME[key-1])
                #sheet = wb[CET_NAME[key-1]]
                #rowNo = 0
                try:
                        driver = init_driver()
                
                        init_Portal(driver)
                        
                        work = 0
                        os.system('cls')
                        lid = f"GMRK08{LOGIN_TYPE[LOGIN_NAME[l]]}{CET_MAP[CET_NAME[key-1]]}"
                        pwd = f"KNR{LOGIN_TYPE[LOGIN_NAME[l]].lower()}@1234"
                        print(f"KNR{LOGIN_TYPE[LOGIN_NAME[l]].lower()}@1234")
                        #pwd = f"GMRK08{LOGIN_TYPE[LOGIN_NAME[l]]}{CET_MAP[CET_NAME[key-1]]}"
                        COLL_CODE = "GMRK08"
                        login_Type = LOGIN_NAME[l]
                        login(driver,lid,pwd,COLL_CODE,login_Type)
                        
                        
                        

                        
                #X = input()
                        while(True):
                                if LOGIN_NAME[l] == 'R':
                                        #driver.execute_script("document.body.style.zoom='150 %'")
                                        
                                        while(True):
                                                try:
                                                        
                                                        wb1 = openpyxl.load_workbook(path_Attempts)
                                                        sheetAT = wb1['ATTEMPTS']
                                                        sheetDT = wb1['DETAILS']
                                                        #print(sheetAT['A1'].value)
                                                        if sheetAT['A1'].value == None:
                                                                sheetAT['A1'] = 0
                                                        TOTAL = int(sheetAT['A1'].value)
                                                        work = 1
                                                        break
                                                except:
                                                        wb1 = openpyxl.Workbook(path_Attempts)
                                                        wb1.create_sheet(index=0, title='ATTEMPTS')
                                                        sheetAT = wb1['ATTEMPTS']
                                                        wb1.create_sheet(index=0, title='DETAILS')
                                                        sheetDT = wb1['DETAILS']
                                                        #sheetAT['A1'] = 0
                                                        #TOTAL = int(sheetAT['A1'].value)
                                                        wb1.save(path_Attempts)
                                                        
                                        
                                        while(True):
                                                print(f'Todays {LOGIN_NAME[l]} Total :{TOTAL}')
                                                TOTAL,ROLL,RANK,name_dsip = reg(TOTAL,driver)
                                                
                                                sheetAT['A1']= TOTAL
                                                sheetDT[f'A{TOTAL}'] =ROLL
                                                sheetDT[f'B{TOTAL}'] =RANK
                                                sheetDT[f'C{TOTAL}'] =name_dsip
                                                sheetDT[f'D{TOTAL}'] =f"COUNTER NO:{(TOTAL%VerOff_num)+1}"
                                                
                                                #print(int(sheetAT['A1'].value))
                                                
                                elif LOGIN_NAME[l] == 'V':
                                        #path_Attempts = f'{path}/{LOGIN_NAME[l]}.xlsx'
                                        while(True):
                                                try:    
                                                        wb1 = openpyxl.load_workbook(path_Attempts)
                                                        sheetAT = wb1['ATTEMPTS']
                                                        if sheetAT['A1'].value == None:
                                                                sheetAT['A1'] = 0
                                                        TOTAL = int(sheetAT['A1'].value)
                                                        work = 1
                                                        break
                                                except:
                                                        wb1 = openpyxl.Workbook(path_Attempts)
                                                        wb1.create_sheet(index=0, title='ATTEMPTS')
                                                        sheetAT = wb1['ATTEMPTS']
                                                        #sheetAT['A1'] = 0
                                                        #TOTAL = int(sheetAT['A1'].value)

                                                        wb1.save(path_Attempts)
                                        while(True):
                                                print(f'Todays {LOGIN_NAME[l]} Total :{TOTAL}')

                                                TOTAL = ver(TOTAL,driver)
                                                
                                                sheetAT['A1']= TOTAL
                                        
                                        
                                
                                
                                
                except Exception as e:
                        print(e)
                        print("Closing....")
                        if work == 1:
                                sheetAT['A1'] = TOTAL
                                print(f"Saving Total No of Candidtaes:{int(sheetAT['A1'].value)}........",end="")                        
                                wb1.save(path_Attempts)
                                print("Saved.")                        
                                print('')
                                work = 0
finally:
                        #print(random.randint(1,999999))
                        #print(TOTAL)
        if work == 1:
                sheetAT['A1'] = TOTAL
                print(f"Saving Total No of Candidtaes:{int(sheetAT['A1'].value)}........",end="")                        
                wb1.save(path_Attempts)
                print("Saved.")                        
                print('')
                        
                        
    
